from datetime import datetime, timedelta
from io import BytesIO
from urllib.parse import quote

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.orm import Session

from app.models import InventoryTransactionRecord, MaterialInventory, ProductDrawing, RawPlateSpecification, ScrapGenerationRecord
from app.services.inventory_summaries import (
    product_summary_rows,
    product_summary_sort_key_map,
    raw_plate_summary_rows,
    raw_plate_summary_sort_key_map,
    scrap_summary_rows,
    scrap_summary_sort_key_map,
)
from app.services.material_matching import scrap_matches_drawing
from app.services.operation_log import record_operation_log
from app.services.drawing_search import drawing_sort_key_map, natural_sort_key
from app.services.list_sorting import sort_records
from app.services.product_outbound_analysis import product_flow_analysis_export_rows
from app.time_utils import china_now


EXPORT_MODULES = {
    "product_inventory": "产品库存",
    "raw_plate_inventory": "板料库存",
    "scrap_inventory": "余料库存",
    "product_transactions": "产品流水",
    "raw_plate_transactions": "板料流水",
    "scrap_transactions": "余料流水",
    "outbound_report": "出库统计",
    "product_outbound_analysis": "产品出库分析",
    "product_catalog": "产品参数清单",
}


def _fmt_time(value) -> str:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else ""


def _fmt_num(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _drawing_version_code(value) -> str:
    return f"A{value or 1}"


def _display_teeth_count(drawing: ProductDrawing) -> object:
    return drawing.teeth_count_text or drawing.teeth_count or ""


def _display_module(drawing: ProductDrawing) -> object:
    return drawing.module_text or drawing.module or ""


def _display_common_normal_length(drawing: ProductDrawing) -> object:
    return drawing.common_normal_length_text or drawing.common_normal_length or ""


def _optional_float(value: str | None) -> float | None:
    try:
        return float(value) if value and value.strip() else None
    except ValueError:
        return None


def _optional_int(value: str | None) -> int | None:
    try:
        return int(value) if value and value.strip() else None
    except ValueError:
        return None


def _float_between_filter(column, value: float, tolerance: float = 0.001):
    return column.between(value - tolerance, value + tolerance)


def _transaction_label(value: str) -> str:
    return {"in": "入库", "out": "出库", "confirm": "确认入库"}.get(value, value)


def _apply_inventory_filters(query, filters: dict, inventory_type: str):
    query = query.filter(MaterialInventory.inventory_type == inventory_type)
    keyword = (filters.get("q") or "").strip()
    if keyword:
        like = f"%{keyword}%"
        query = query.filter(
            (MaterialInventory.material_code.ilike(like))
            | (MaterialInventory.material.ilike(like))
            | (MaterialInventory.location.ilike(like))
            | (MaterialInventory.usable_size.ilike(like))
            | (MaterialInventory.source_product_code.ilike(like))
            | (MaterialInventory.paper_material.ilike(like))
        )
    material = (filters.get("material") or "").strip()
    if material:
        query = query.filter(MaterialInventory.material.ilike(f"%{material}%"))
    thickness = _optional_float(filters.get("thickness"))
    if thickness is not None:
        query = query.filter(MaterialInventory.thickness == thickness)
    location = (filters.get("location") or "").strip()
    if location:
        query = query.filter(MaterialInventory.location.ilike(f"%{location}%"))
    if inventory_type == "raw_plate":
        length = _optional_float(filters.get("length"))
        if length is not None:
            query = query.filter(MaterialInventory.length == length)
        width = _optional_float(filters.get("width"))
        if width is not None:
            query = query.filter(MaterialInventory.width == width)
    return query


def _sorted_summary_rows(rows: list[dict], filters: dict, key_map: dict, default_key) -> list[dict]:
    sorted_rows, selected_field, _ = sort_records(
        rows,
        filters.get("sort_by") or "",
        filters.get("sort_dir") or "",
        key_map,
    )
    if not selected_field:
        sorted_rows.sort(key=default_key)
    return sorted_rows


def _joined_numbers(values: set) -> str:
    return " / ".join(_fmt_num(value) for value in sorted(value for value in values if value is not None))


def _scrap_inventory_export_records(db: Session, filters: dict) -> tuple[list[ScrapGenerationRecord], dict[int, MaterialInventory]]:
    records = db.query(ScrapGenerationRecord).order_by(ScrapGenerationRecord.registered_at.desc()).all()
    scrap_ids = [record.scrap_inventory_id for record in records if record.scrap_inventory_id]
    scrap_map = {
        item.id: item
        for item in db.query(MaterialInventory).filter(MaterialInventory.id.in_(scrap_ids)).all()
    } if scrap_ids else {}
    keyword = (filters.get("source_product_code") or "").strip().lower()
    material = (filters.get("material") or "").strip()
    thickness = _optional_float(filters.get("thickness"))
    required_diameter = _optional_float(filters.get("required_diameter"))
    location = (filters.get("location") or "").strip()
    drawing_text = str(filters.get("drawing_id") or "").strip()
    drawing = db.get(ProductDrawing, int(drawing_text)) if drawing_text.isdigit() else None
    filtered = []
    for record in records:
        item = scrap_map.get(record.scrap_inventory_id)
        if not item or item.status != "available":
            continue
        searchable = " ".join(str(value or "") for value in (record.source_product_code, item.material, item.usable_size, item.location)).lower()
        if keyword and keyword not in searchable:
            continue
        if material and material not in item.material:
            continue
        if thickness is not None and item.thickness != thickness:
            continue
        if required_diameter is not None and (item.diameter is None or item.diameter < required_diameter):
            continue
        if drawing and not scrap_matches_drawing(item, drawing):
            continue
        if location and location not in (item.location or ""):
            continue
        filtered.append(record)
    return filtered, scrap_map


def _date_range(filters: dict) -> tuple[datetime | None, datetime | None]:
    start_value = (filters.get("start_date") or "").strip()
    end_value = (filters.get("end_date") or "").strip()
    start = datetime.strptime(start_value, "%Y-%m-%d") if start_value else None
    end = datetime.strptime(end_value, "%Y-%m-%d") + timedelta(days=1) if end_value else None
    return start, end


def _transaction_rows(db: Session, inventory_type: str, filters: dict) -> tuple[list[str], list[list[object]]]:
    query = db.query(InventoryTransactionRecord).order_by(InventoryTransactionRecord.created_at.desc())
    start, end = _date_range(filters)
    if start:
        query = query.filter(InventoryTransactionRecord.created_at >= start)
    if end:
        query = query.filter(InventoryTransactionRecord.created_at < end)
    records = query.all()
    inventory_ids = [record.inventory_id for record in records]
    inventory_map = {item.id: item for item in db.query(MaterialInventory).filter(MaterialInventory.id.in_(inventory_ids)).all()} if inventory_ids else {}
    product_code = (filters.get("product_code") or filters.get("q") or "").strip()
    rows = []
    for record in records:
        item = inventory_map.get(record.inventory_id)
        if not item or item.inventory_type != inventory_type:
            continue
        code = item.material_code or item.source_product_code or item.material or ""
        if product_code and product_code not in code:
            continue
        rows.append([
            record.id,
            _transaction_label(record.transaction_type),
            code,
            record.quantity,
            record.before_quantity,
            record.after_quantity,
            record.customer_name or "",
            record.operator_name or "",
            record.remark or "",
            _fmt_time(record.created_at),
        ])
    return ["流水号", "类型", "型号/规格", "数量", "操作前库存", "操作后库存", "客户/去向", "操作人", "备注", "创建时间"], rows


def _outbound_item_values(item: MaterialInventory) -> tuple[str, str, str, str]:
    if item.inventory_type == "product":
        return (
            item.material_code or item.source_product_code or "-",
            item.material or "-",
            f"厚度 {_fmt_num(item.thickness)}" if item.thickness is not None else "-",
            item.location or "-",
        )
    if item.inventory_type == "raw_plate":
        return (
            item.material or "-",
            item.usable_size or f"{_fmt_num(item.length) or '-'}×{_fmt_num(item.width) or '-'}×{_fmt_num(item.thickness) or '-'}mm",
            f"长 {_fmt_num(item.length) or '-'}｜宽 {_fmt_num(item.width) or '-'}｜厚 {_fmt_num(item.thickness) or '-'}",
            item.location or "-",
        )
    return (
        item.material or "-",
        item.usable_size or "-",
        f"厚度 {_fmt_num(item.thickness)}｜直径 {_fmt_num(item.diameter)}" if item.diameter is not None else f"厚度 {_fmt_num(item.thickness)}",
        item.location or "-",
    )


def _outbound_report_rows(db: Session, filters: dict) -> tuple[list[str], list[list[object]]]:
    start, end = _report_range(filters)
    records = db.query(InventoryTransactionRecord).filter(
        InventoryTransactionRecord.transaction_type == "out",
        InventoryTransactionRecord.reversed_transaction_id.is_(None),
        InventoryTransactionRecord.created_at >= start,
        InventoryTransactionRecord.created_at < end,
    ).order_by(InventoryTransactionRecord.created_at.desc(), InventoryTransactionRecord.id.desc()).all()
    inventory_ids = [record.inventory_id for record in records]
    inventory_map = {item.id: item for item in db.query(MaterialInventory).filter(MaterialInventory.id.in_(inventory_ids)).all()} if inventory_ids else {}
    range_label = f"{start.strftime('%Y-%m-%d')} 至 {(end - timedelta(days=1)).strftime('%Y-%m-%d')}"
    rows = []
    for record in records:
        item = inventory_map.get(record.inventory_id)
        if not item:
            continue
        type_label = {"product": "产品", "raw_plate": "板料", "scrap": "余料"}.get(item.inventory_type, item.inventory_type)
        code, name, spec, location = _outbound_item_values(item)
        rows.append([
            record.id,
            _fmt_time(record.created_at),
            type_label,
            code,
            name,
            spec,
            location,
            record.customer_name or "",
            record.quantity,
            record.operator_name or "",
            record.remark or "",
            range_label,
        ])
    return ["流水号", "出库时间", "类型", "型号/规格", "材质/名称", "规格", "库位", "客户/去向", "出库数量", "操作人", "备注", "时间范围"], rows


def _report_range(filters: dict) -> tuple[datetime, datetime]:
    start_value = (filters.get("start_date") or "").strip()
    end_value = (filters.get("end_date") or "").strip()
    if start_value and end_value:
        return datetime.strptime(start_value, "%Y-%m-%d"), datetime.strptime(end_value, "%Y-%m-%d") + timedelta(days=1)
    now = china_now()
    period = (filters.get("period") or "day").strip()
    if period == "month":
        return datetime(now.year, now.month, 1), now + timedelta(days=1)
    if period == "year":
        return datetime(now.year, 1, 1), now + timedelta(days=1)
    return datetime(now.year, now.month, now.day), now + timedelta(days=1)


def _apply_drawing_filters(query, filters: dict):
    keyword = (filters.get("q") or "").strip()
    if keyword:
        like = f"%{keyword}%"
        keyword_filter = (
            (ProductDrawing.product_code.ilike(like))
            | (ProductDrawing.product_name.ilike(like))
            | (ProductDrawing.product_category.ilike(like))
            | (ProductDrawing.remark.ilike(like))
            | (ProductDrawing.material.ilike(like))
            | (ProductDrawing.tooth_type.ilike(like))
            | (ProductDrawing.teeth_count_text.ilike(like))
            | (ProductDrawing.module_text.ilike(like))
            | (ProductDrawing.common_normal_length_text.ilike(like))
        )
        query = query.filter(keyword_filter)
    product_category = (filters.get("product_category") or "").strip()
    if product_category:
        query = query.filter(ProductDrawing.product_category.ilike(f"%{product_category}%"))
    material = (filters.get("material") or "").strip()
    if material:
        query = query.filter(ProductDrawing.material.ilike(f"%{material}%"))
    thickness = _optional_float(filters.get("thickness"))
    if thickness is not None:
        query = query.filter(
            _float_between_filter(ProductDrawing.thickness, thickness)
            | _float_between_filter(ProductDrawing.product_thickness, thickness)
            | _float_between_filter(ProductDrawing.plate_thickness, thickness)
        )
    product_thickness = _optional_float(filters.get("product_thickness"))
    if product_thickness is not None:
        query = query.filter(_float_between_filter(ProductDrawing.product_thickness, product_thickness))
    plate_thickness = _optional_float(filters.get("plate_thickness"))
    if plate_thickness is not None:
        query = query.filter(_float_between_filter(ProductDrawing.plate_thickness, plate_thickness))
    outer_diameter = _optional_float(filters.get("outer_diameter"))
    if outer_diameter is not None:
        query = query.filter(_float_between_filter(ProductDrawing.max_outer_diameter, outer_diameter))
    inner_diameter = _optional_float(filters.get("inner_diameter"))
    if inner_diameter is not None:
        query = query.filter(_float_between_filter(ProductDrawing.min_inner_diameter, inner_diameter))
    teeth_count_text = (filters.get("teeth_count") or "").strip()
    if teeth_count_text:
        teeth_count = _optional_int(teeth_count_text)
        like = f"%{teeth_count_text}%"
        if teeth_count is not None:
            query = query.filter((ProductDrawing.teeth_count == teeth_count) | ProductDrawing.teeth_count_text.ilike(like))
        else:
            query = query.filter((ProductDrawing.teeth_count_text.ilike(like)) | ProductDrawing.tooth_type.ilike(like))
    module_text = (filters.get("module") or "").strip()
    if module_text:
        module = _optional_float(module_text)
        like = f"%{module_text}%"
        if module is not None:
            query = query.filter(_float_between_filter(ProductDrawing.module, module) | ProductDrawing.module_text.ilike(like))
        else:
            query = query.filter(ProductDrawing.module_text.ilike(like))
    pressure_angle = _optional_float(filters.get("pressure_angle"))
    if pressure_angle is not None:
        query = query.filter(_float_between_filter(ProductDrawing.pressure_angle, pressure_angle))
    common_normal_length_text = (filters.get("common_normal_length") or "").strip()
    if common_normal_length_text:
        common_normal_length = _optional_float(common_normal_length_text)
        like = f"%{common_normal_length_text}%"
        if common_normal_length is not None:
            query = query.filter(_float_between_filter(ProductDrawing.common_normal_length, common_normal_length) | ProductDrawing.common_normal_length_text.ilike(like))
        else:
            query = query.filter(ProductDrawing.common_normal_length_text.ilike(like))
    pin_diameter = _optional_float(filters.get("pin_diameter"))
    if pin_diameter is not None:
        query = query.filter(_float_between_filter(ProductDrawing.pin_diameter, pin_diameter))
    pin_span = _optional_float(filters.get("pin_span"))
    if pin_span is not None:
        query = query.filter(_float_between_filter(ProductDrawing.pin_span, pin_span))
    return query


def _product_catalog_rows(db: Session, filters: dict) -> tuple[list[str], list[list[object]]]:
    query = db.query(ProductDrawing).filter(ProductDrawing.confirmed == 1, ProductDrawing.is_active == 1)
    drawings = sorted(
        _apply_drawing_filters(query, filters).all(),
        key=lambda drawing: (natural_sort_key(drawing.product_code), -(drawing.version or 1)),
    )
    drawings, _, _ = sort_records(
        drawings,
        filters.get("sort_by") or "",
        filters.get("sort_dir") or "",
        drawing_sort_key_map(),
    )
    headings = [
        "产品分类",
        "产品型号",
        "产品名称",
        "材质",
        "产品厚度",
        "钢板厚度",
        "外径",
        "内径",
        "齿型",
        "齿数",
        "模数",
        "压力角",
        "变位系数",
        "跨齿数",
        "公法线长度",
        "量棒直径",
        "棒间距",
        "中心余料尺寸",
        "备注",
        "版本",
        "更新时间",
    ]
    rows = [
        [
            drawing.product_category or "",
            drawing.product_code or "",
            drawing.product_name or "",
            drawing.material or "",
            _fmt_num(drawing.product_thickness),
            _fmt_num(drawing.plate_thickness),
            _fmt_num(drawing.max_outer_diameter),
            _fmt_num(drawing.min_inner_diameter),
            drawing.tooth_type or "",
            _display_teeth_count(drawing),
            _display_module(drawing),
            _fmt_num(drawing.pressure_angle),
            _fmt_num(drawing.profile_shift_coefficient),
            drawing.span_teeth_count or "",
            _display_common_normal_length(drawing),
            _fmt_num(drawing.pin_diameter),
            _fmt_num(drawing.pin_span),
            drawing.expected_scrap_size or "",
            drawing.remark or "",
            _drawing_version_code(drawing.version),
            _fmt_time(drawing.updated_at),
        ]
        for drawing in drawings
    ]
    return headings, rows


def build_export_rows(module: str, filters: dict, db: Session) -> tuple[str, list[str], list[list[object]]]:
    if module not in EXPORT_MODULES:
        raise HTTPException(status_code=404, detail="导出模块不存在")
    if module == "product_catalog":
        return EXPORT_MODULES[module], *_product_catalog_rows(db, filters)
    if module == "product_inventory":
        items = _apply_inventory_filters(db.query(MaterialInventory), filters, "product").order_by(MaterialInventory.created_at.desc()).all()
        groups = _sorted_summary_rows(
            product_summary_rows(items),
            filters,
            product_summary_sort_key_map(),
            lambda row: natural_sort_key(row["code"]),
        )
        rows = [[group["code"], group["quantity"], group["material"], _joined_numbers(group["product_thicknesses"]), _joined_numbers(group["plate_thicknesses"]), " / ".join(sorted(group["paper_materials"])), group["batch_count"], " / ".join(sorted(group["locations"])), _fmt_time(group["latest"])] for group in groups]
        return EXPORT_MODULES[module], ["产品型号", "库存数量", "材质", "总成品厚度", "钢板厚度", "纸材质", "批次数", "库位", "最近更新时间"], rows
    if module == "raw_plate_inventory":
        items = _apply_inventory_filters(db.query(MaterialInventory), filters, "raw_plate").order_by(MaterialInventory.created_at.desc()).all()
        spec_names = {
            (spec.material, spec.length, spec.width, spec.thickness): spec.spec_name
            for spec in db.query(RawPlateSpecification).filter(RawPlateSpecification.is_active == 1).all()
        }
        groups = _sorted_summary_rows(
            raw_plate_summary_rows(items, spec_names),
            filters,
            raw_plate_summary_sort_key_map(),
            lambda row: (str(row["material"]), row["thickness"] or 0, row["length"] or 0, row["width"] or 0),
        )
        rows = [[group["spec_name"], group["material"], _fmt_num(group["length"]), _fmt_num(group["width"]), _fmt_num(group["thickness"]), group["quantity"], group["batch_count"], " / ".join(sorted(group["locations"])), _fmt_time(group["latest"])] for group in groups]
        return EXPORT_MODULES[module], ["规格", "材质", "长度", "宽度", "厚度", "总块数", "批次数", "库位", "最近更新时间"], rows
    if module == "scrap_inventory":
        records, scrap_map = _scrap_inventory_export_records(db, filters)
        groups = _sorted_summary_rows(
            scrap_summary_rows(records, scrap_map),
            filters,
            scrap_summary_sort_key_map(),
            lambda row: (str(row["material"]), row["thickness"] or 0, str(row["usable_size"])),
        )
        rows = [[group["material"], _fmt_num(group["thickness"]), group["usable_size"], group["quantity"], group["batch_count"], " / ".join(sorted(group["locations"])), _fmt_time(group["latest"])] for group in groups]
        return EXPORT_MODULES[module], ["材质", "厚度", "可用尺寸", "总数量", "批次数", "库位", "最近更新时间"], rows
    if module == "product_transactions":
        return EXPORT_MODULES[module], *_transaction_rows(db, "product", filters)
    if module == "raw_plate_transactions":
        return EXPORT_MODULES[module], *_transaction_rows(db, "raw_plate", filters)
    if module == "scrap_transactions":
        return EXPORT_MODULES[module], *_transaction_rows(db, "scrap", filters)
    if module == "product_outbound_analysis":
        title = "产品入库分析" if (filters.get("flow_type") or "").strip() == "in" else EXPORT_MODULES[module]
        return title, *product_flow_analysis_export_rows(db, filters)
    return EXPORT_MODULES[module], *_outbound_report_rows(db, filters)


def make_workbook_bytes(title: str, headings: list[str], rows: list[list[object]]) -> BytesIO:
    workbook = Workbook(write_only=False)
    sheet = workbook.active
    sheet.title = title[:31]
    sheet.append(headings)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 36)
        sheet.column_dimensions[column_cells[0].column_letter].width = width
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def export_filename(title: str) -> str:
    return f"{title}_{china_now().strftime('%Y%m%d_%H%M%S')}.xlsx"


def content_disposition(filename: str) -> str:
    return f"attachment; filename*=UTF-8''{quote(filename)}"


def log_export(module: str, filters: dict, row_count: int, db: Session) -> None:
    record_operation_log(
        db,
        "excel_export",
        module,
        None,
        None,
        f"导出{EXPORT_MODULES.get(module, module)}，共{row_count}行",
        after_data={"module": module, "filters": filters, "row_count": row_count, "export_time": china_now().isoformat()},
    )
